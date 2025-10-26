# extract data from multiple CRA files in several excel files into  a single csv file
from concurrent.futures import ProcessPoolExecutor
from openpyxl import load_workbook
import csv
import time



# configuration variables
FOLDER_PREFIX = 'CRA'
FILE_EXTENSION = '.xlsm'
FILE_NAME = 'TEST/flat/CRA Ann‚e 2025_LEROY Julien.xlsm'
WORKSHEET_NAME = 'MàJ CRA'
BASE_DIR = '/Users/jleroy/Documents/dev/CRA/TEST/CRA TS'  # set to None to use the script directory

# main function
def main():
    print("Extraction des données des fichiers CRA vers des fichiers CSV")
    print("Prenez un café, cela peut prendre un certain temps... (4h00 pour 100 fichiers environ sur un MacBook Pro 2020 M1)")

    # démarrer le chrono global
    start_time = time.perf_counter()

    # get the list of CRA files
    file_list = get_file_list()
    # TEST: ne traiter qu'un seul fichier
    # file_list =  [file_list[0]]

    # process each file
    with ProcessPoolExecutor() as pool:
        # for data_file in file_list:
        for wb in pool.map(process_file, file_list):
            pass

    # print the total elapsed time
    total_elapsed = time.perf_counter() - start_time
    print(f"Temps total : {total_elapsed:.2f}s")
    return

# process a single CRA file
def process_file(data_file=None):
    # load the entire workbook
    file_start = time.perf_counter()
    print('Processing file:', data_file)
    wb = load_workbook(filename=data_file, read_only=True, data_only=True, keep_links=False)
    ws = wb[WORKSHEET_NAME]

    # prepare the header
    header = []
    header.append("annee")
    header.append("collaborateur")
    header.append("matricule")
    header.append("date_saisie")
    header.append("id_categorie")
    header.append("libelle_categorie")
    header.append("detail_activite")
    header.append("nb_jour")
    header_string=";".join(header)
    # print(header_string)

    rowdata = []
    # année
    rowdata.append(str(ws['B1'].value))
    # collaborateur
    rowdata.append(ws['B2'].value)
    # matricule
    rowdata.append(str(ws['B3'].value))
    # date_saisie

    # foreach week column
    rowlist=[]
    rowlist=process_week(ws,rowdata)
    write_csv(filename=data_file+".csv", header=header, data=rowlist)
    wb.close()
    file_elapsed = time.perf_counter() - file_start
    print(f"Temps traitement {data_file} : {file_elapsed:.2f}s")
    return

def process_week(ws=None,rowdata=[]):
    rowdata_string = []
    # process category
    for col in range(4, 4+365):
    # TODO: manage leap years
    # TODO: count the number of holidays per week on row 7
        day_value = ws.cell(row=5, column=col).value
        if day_value is None:
            # print('  No day value at column:', col)
            continue
        if day_value.date().weekday() == 0 or (day_value.date().day == 1 and day_value.date().weekday() != 5 and day_value.date().weekday() != 6):
            print('  Processing week:', col-3)
            # print (f"jour: {day_value.date()}/jour de la semaine: {day_value.date().weekday()}/jour du mois: {day_value.date().day}")
            processed_categories=process_category(ws,col,rowdata)
            # print('  Processed categories:', processed_categories)
            rowdata_string += processed_categories
            # rowdata_string.append(";".join(rowdata))
            # print(rowdata_string)
        else:
            continue
    return rowdata_string

def process_category(ws=None,col=4,rowdata=[]):
    rowstring = []
    row_index = 8
    category_row_index = row_index
    while (current_category := ws.cell(row=row_index, column=2).value) is not None:
        if current_category != "OK":
            category_row_index=row_index
            # print('    Processing category:', current_category)
        else:
            # print('    Processing activity:', ws.cell(row=row_index, column=3).value)
            nb_jour = ws.cell(row=row_index, column=col)
            if nb_jour.value is not None:
                # date_saisie
                date_saisie = ws.cell(row=5, column=col).value
                rowdata.append(f"{date_saisie.date()}")
                # id_categorie
                rowdata.append(str(ws.cell(row=category_row_index, column=2).value))
                # Libelle_categorie
                rowdata.append(str(ws.cell(row=category_row_index, column=3).value))
                # detail_activite
                rowdata.append(str(ws.cell(row=row_index, column=3).value))
                # nb_jour
                if isinstance(nb_jour.value, int) or isinstance(nb_jour.value, float):
                    # print('      Processing activity:', offset)
                    # print ('     ',str(ws.cell(row=line_index, column=3).value))
                    rowdata.append(str(nb_jour.value))
                rowstring.append(rowdata)
        row_index += 1
    return rowstring

def get_file_list(base_dir=None, prefix=FOLDER_PREFIX, extension=FILE_EXTENSION):
    # This function return the list of CRA files
    # which the name ends by ".xlsm"
    # that belongs in the folder which the name starts by "CRA"
    import os
    # par défaut on cherche depuis le dossier du script
    if base_dir is None:
        base_dir = os.path.abspath(os.path.dirname(__file__))

    matches = []
    for root, dirs, files in os.walk(base_dir):
        # si le nom du dossier commence par "CRA"
        if os.path.basename(root).startswith(prefix):
            for fname in files:
                if fname.lower().endswith(extension.lower()):
                    matches.append(os.path.join(root, fname))
    matches.sort()
    # print(matches)
    return matches

def write_csv(filename="output.csv", header=[], data=[]):
    with open(filename, mode='w', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=';')
        if header:
            csvwriter.writerow(header)
        if data:
            csvwriter.writerows(data)
        # for row in data:
            # csvwriter.writerows(row)
        csvfile.close()
    return

if __name__ == "__main__":
    main()
